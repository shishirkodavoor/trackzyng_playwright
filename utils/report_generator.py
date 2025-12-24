"""Generate Excel report from pytest results."""
import json
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.test_case_mapping import get_test_case_id, TEST_CASE_MAPPING

def parse_allure_results(allure_results_dir: Path):
    """Parse Allure results JSON files."""
    results = []
    
    if not allure_results_dir.exists():
        return results
    
    for result_file in allure_results_dir.glob("*.json"):
        try:
            with open(result_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
                test_name = data.get('name', 'Unknown Test')
                status = data.get('status', 'unknown')
                status_details = data.get('statusDetails', {})
                start_time = data.get('start', 0)
                stop_time = data.get('stop', 0)
                duration = (stop_time - start_time) / 1000 if stop_time and start_time else 0
                
                # Extract test case ID
                test_case_id = get_test_case_id(test_name)
                
                # Get error message if failed
                error_message = ""
                if status == 'failed' and status_details:
                    error_message = status_details.get('message', '')
                    trace = status_details.get('trace', '')
                    if trace:
                        error_message += f"\n{trace[:200]}"
                
                results.append({
                    'test_case_id': test_case_id,
                    'test_name': test_name,
                    'status': status.upper() if status else 'UNKNOWN',
                    'duration': duration,
                    'error_message': error_message,
                    'start_time': datetime.fromtimestamp(start_time / 1000).strftime('%Y-%m-%d %H:%M:%S') if start_time else '',
                })
        except Exception as e:
            print(f"Error parsing {result_file}: {e}")
    
    return results

def generate_excel_report(output_path: Path, allure_results_dir: Path):
    """Generate Excel report with test results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Headers
    headers = [
        "Test Case ID",
        "Test Scenario",
        "Test Steps",
        "Test Data",
        "Precondition",
        "Expected Output",
        "Actual Output",
        "Status",
        "Remarks",
        "Tested By",
        "Tested Date",
        "Duration (seconds)"
    ]
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Parse results
    results = parse_allure_results(allure_results_dir)
    
    # If no results, create empty row with message
    if not results:
        row = 2
        ws.cell(row=row, column=1, value="No test results found")
        ws.cell(row=row, column=8, value="NO DATA")
        ws.merge_cells(f'A{row}:L{row}')
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = border
    else:
        # Write test results
        for row_num, result in enumerate(results, 2):
            test_name = result['test_name']
            status = result['status']
            error_message = result.get('error_message', '')
            duration = result.get('duration', 0)
            test_case_id = result.get('test_case_id', 'TC_UNKNOWN')
            tested_date = result.get('start_time', datetime.now().strftime('%Y-%m-%d'))
            
            # Map status to Pass/Fail
            status_display = "Pass" if status == "PASSED" else "Fail" if status == "FAILED" else status
            
            # Determine test scenario from test name
            test_scenario = test_name.replace('test_', '').replace('_', ' ').title()
            
            # Fill row data
            ws.cell(row=row_num, column=1, value=test_case_id).border = border
            ws.cell(row=row_num, column=2, value=test_scenario).border = border
            ws.cell(row=row_num, column=3, value="1. Open application\n2. Navigate to page\n3. Perform actions\n4. Verify results").border = border
            ws.cell(row=row_num, column=4, value="Test data as per test scenario").border = border
            ws.cell(row=row_num, column=5, value="Application is accessible").border = border
            ws.cell(row=row_num, column=6, value=f"Test should {status_display.lower()}").border = border
            ws.cell(row=row_num, column=7, value=error_message[:500] if error_message else status_display).border = border
            ws.cell(row=row_num, column=8, value=status_display).border = border
            
            # Color code status column
            status_cell = ws.cell(row=row_num, column=8)
            if status_display == "Pass":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status_display == "Fail":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws.cell(row=row_num, column=9, value="").border = border  # Remarks
            ws.cell(row=row_num, column=10, value="Automated").border = border  # Tested By
            ws.cell(row=row_num, column=11, value=tested_date).border = border  # Tested Date
            ws.cell(row=row_num, column=12, value=f"{duration:.2f}").border = border  # Duration
            
            # Set alignment
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(header)
        column = get_column_letter(col_num)
        
        for row in ws[column]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel report generated: {output_path}")

if __name__ == "__main__":
    # Example usage
    base_dir = Path(__file__).parent.parent
    allure_dir = base_dir / "reports" / "allure-results"
    output_file = base_dir / "reports" / f"Test_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    generate_excel_report(output_file, allure_dir)

